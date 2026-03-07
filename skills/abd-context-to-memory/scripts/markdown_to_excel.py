#!/usr/bin/env python3
"""
Generic markdown → Excel. Parses headings, tables, paragraphs; writes to a new workbook.

Usage:
  python markdown_to_excel.py <input.md> [output.xlsx]
  python markdown_to_excel.py --file <input.md> [--out <output.xlsx>]

Creates a workbook with:
- Top-level headings (# ) as section headers
- Tables as Excel tables (header row + data rows)
- Paragraphs as rows (heading in col A, content in col B)

Requires: pip install openpyxl
"""
import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

BULLET = "\u2022"


def parse_markdown(content: str) -> list[dict]:
    """Parse markdown into blocks: {type: heading|table|paragraph, level?, text, rows?}."""
    blocks = []
    lines = content.split("\n")
    i = 0

    while i < len(lines):
        line = lines[i]
        # Heading
        h_match = re.match(r"^(#{1,6})\s+(.+)$", line)
        if h_match:
            level = len(h_match.group(1))
            blocks.append({"type": "heading", "level": level, "text": h_match.group(2).strip()})
            i += 1
            continue

        # Table
        if re.match(r"^\s*\|.+\|\s*$", line):
            rows = []
            while i < len(lines) and re.match(r"^\s*\|.+\|\s*$", lines[i]):
                row_line = lines[i]
                cells = [c.strip() for c in row_line.split("|")[1:-1]]
                if cells and not all(re.match(r"^[-:\s]+$", c) for c in cells):
                    rows.append(cells)
                i += 1
            if rows:
                blocks.append({"type": "table", "rows": rows})
            continue

        # Paragraph (non-empty, not list-only)
        para_lines = []
        while i < len(lines) and lines[i].strip() and not lines[i].startswith("#") and not re.match(r"^\s*\|", lines[i]):
            para_lines.append(lines[i])
            i += 1
        if para_lines:
            blocks.append({"type": "paragraph", "text": "\n".join(para_lines).strip()})
            continue

        i += 1

    return blocks


def md_to_excel_plain(text: str) -> str:
    """Convert basic markdown to plain text for Excel: **bold** → plain, - → •."""
    t = re.sub(r"^(\s*)- ", r"\1" + BULLET + " ", text, flags=re.MULTILINE)
    t = re.sub(r"\*\*([^*]+)\*\*", r"\1", t)
    t = re.sub(r"\*([^*]+)\*", r"\1", t)
    return t


def write_blocks_to_excel(blocks: list[dict], ws) -> None:
    """Write parsed blocks to worksheet."""
    row = 1
    current_heading = ""

    for blk in blocks:
        if blk["type"] == "heading":
            current_heading = blk["text"]
            level = blk.get("level", 1)
            cell = ws.cell(row, 1, current_heading)
            cell.font = Font(bold=True, size=14 if level <= 2 else 12)
            row += 1

        elif blk["type"] == "table":
            for r_idx, r_cells in enumerate(blk["rows"]):
                for c_idx, val in enumerate(r_cells):
                    cell = ws.cell(row + r_idx, c_idx + 1, md_to_excel_plain(val))
                    if r_idx == 0:
                        cell.font = Font(bold=True)
            row += len(blk["rows"]) + 1

        elif blk["type"] == "paragraph":
            if current_heading:
                ws.cell(row, 1, current_heading).font = Font(bold=True)
            cell = ws.cell(row, 2, md_to_excel_plain(blk["text"]))
            cell.alignment = Alignment(wrap_text=True)
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert markdown to Excel")
    parser.add_argument("input", nargs="?", help="Input markdown file")
    parser.add_argument("output", nargs="?", help="Output Excel file (default: input stem + .xlsx)")
    parser.add_argument("--file", "-f", help="Input markdown file")
    parser.add_argument("--out", "-o", help="Output Excel file")
    args = parser.parse_args()

    md_path = args.file or args.input
    if not md_path:
        parser.print_help()
        return 1

    md_path = Path(md_path).resolve()
    if not md_path.exists():
        print(f"ERROR: File not found: {md_path}")
        return 1

    out_path = args.out or args.output
    if not out_path:
        out_path = md_path.with_suffix(".xlsx")
    else:
        out_path = Path(out_path).resolve()

    content = md_path.read_text(encoding="utf-8")
    blocks = parse_markdown(content)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content"
    write_blocks_to_excel(blocks, ws)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
